"""
workbook.py — Fill the RR calculator template with extracted data and engine outputs.
Uses approach (b): write both inputs AND engine-computed outputs as values.
No LibreOffice dependency; output is an xlsx the reviewer can open directly.
"""
import base64
import io
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates",
                              "rr_calculator_template.xlsx")

_GREEN = "FF92D050"
_ORANGE = "FFFFC000"
_RED = "FFFF0000"

_COLOUR_MAP = {
    "10%": _GREEN,
    "5%": _ORANGE,
    "Flat": _GREEN,
    "Waiver_incl": _RED,
    "Waiver_noCHB": _GREEN,
}


def _money_str(v: float) -> str:
    v = round(v)
    return f"-${abs(v):,}" if v < 0 else f"${v:,}"


def fill_workbook(merchant: dict, months: list[dict], engine: dict,
                  flat_amount: float, window_months: list[str]) -> bytes:
    """
    Build and return the filled .xlsx as bytes.
    If template exists, uses it as base; otherwise creates a structured workbook.
    merchant: the merchant input dict
    months: [{"label": YYYY-MM, "sales": float, "refunds": float, "chargebacks": float}]
    engine: output of calc.rr_engine()
    """
    if not HAS_OPENPYXL:
        return b""

    template_path = TEMPLATE_PATH
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RR Assessment"

    # --- Header block ---
    ws["A1"] = "Rolling Reserve Assessment"
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"] = ""
    row = 4
    for label, key in [
        ("Guesty ID", "guesty_id"),
        ("Guesty Name", "guesty_name"),
        ("GuestyPay", "guestypay"),
        ("Region", "region"),
        ("Bank", "bank"),
        ("Segment", "segment"),
        ("Avg. MRR", "avg_mrr"),
        ("Active Listings", "active_listings"),
        ("Requested Change", "requested_change"),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=merchant.get(key, ""))
        row += 1

    row += 1

    # --- Monthly data table ---
    headers = ["Month", "Sales $", "Chargebacks $", "CHB Rate %",
               "Refunds $", "Refund Rate %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
    row += 1

    from app.calc import monthly_rates
    rates = monthly_rates({m["label"]: m for m in months})

    for m in months:
        mk = m["label"]
        r = rates.get(mk, {})
        ws.cell(row=row, column=1, value=mk)
        ws.cell(row=row, column=2, value=m["sales"])
        ws.cell(row=row, column=3, value=m["chargebacks"])
        chb_rate = r.get("chb_rate") or 0.0
        ws.cell(row=row, column=4, value=round(chb_rate * 100, 4))
        ws.cell(row=row, column=5, value=m["refunds"])
        ref_rate = r.get("refund_rate") or 0.0
        ws.cell(row=row, column=6, value=round(ref_rate * 100, 4))
        row += 1

    row += 1
# --- Fill the template's embedded month table (rows 24-29, cols C/D/E) ---
    # Overwrites the stale Dec/Jan/Feb headers with the real months and figures.
    grid_cols = ["C", "D", "E", "F", "G"]
    for i, m in enumerate(months[:5]):
        col = grid_cols[i]
        from datetime import datetime as _dt
        ws[f"{col}24"] = _dt.strptime(m["label"], "%Y-%m").strftime("%b %Y")  # Month header e.g. "Apr 2026"          # Month header
        ws[f"{col}25"] = m["sales"]          # Sales Volume $
        ws[f"{col}26"] = m["chargebacks"]    # Chargeback Volume $
        ws[f"{col}29"] = m["refunds"]        # Refund volume $
    # --- Summary ---
    vol_row = row
    ws.cell(row=row, column=1, value="Volume 90d")
    ws.cell(row=row, column=2, value=engine.get("_volume_90d", ""))
    row += 1
    # Column D (Volume Last 90 Days) for all 5 scenario rows -> computed Volume 90d cell
    for r in range(2, 8):
        ws.cell(row=r, column=4, value=f"=$B${vol_row}")
    ws.cell(row=row, column=1, value="Highest CHB Rate")
    ws.cell(row=row, column=2, value=engine.get("_highest_chb_rate", ""))
    row += 1
    ref_row = row
    ws.cell(row=row, column=1, value="Highest Refund Rate")
    ws.cell(row=row, column=2, value=engine.get("_highest_refund_rate", ""))
    row += 1
    for r in range(2, 8):
        ws.cell(row=r, column=7, value=f"=$B${ref_row}")
    ws.cell(row=row, column=1, value="Estimated CHB")
    ws.cell(row=row, column=2, value=engine["estimated_chb"])
    row += 1
    ws.cell(row=row, column=1, value="Estimated Refund")
    ws.cell(row=row, column=2, value=engine["estimated_refund"])
    row += 1
    ws.cell(row=row, column=1, value="CNL")
    ws.cell(row=row, column=2, value=engine["CNL"])
    row += 2

    # --- Exposure scenarios ---
    scenario_headers = ["", "RR Option", "Days", "Exposure $"]
    for col, h in enumerate(scenario_headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1

    exp = engine["EXP"]
    scenarios = [
        ("10%", "10%", "90", exp["10%"]),
        ("5%", "5%", "90", exp["5%"]),
        ("Flat", f"FLAT ${int(flat_amount/1000)}K", str(int(flat_amount)), exp["Flat"]),
        ("Waiver_incl", "Waiver incl. CHBs+RFU", "", exp["Waiver_incl"]),
        ("Waiver_noCHB", "Waiver no CHBs, RFU<13%", "", exp["Waiver_noCHB"]),
    ]
    for key, label, days, exposure in scenarios:
        colour = _COLOUR_MAP.get(key, "FFFFFFFF")
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        c1 = ws.cell(row=row, column=1, value=key)
        c1.fill = fill
        ws.cell(row=row, column=2, value=label)
        _align = Alignment(horizontal="center")
        if days:
            dc = ws.cell(row=row, column=3, value=days)
            dc.alignment = _align
            ec = ws.cell(row=row, column=4, value=_money_str(exposure))
            ec.alignment = _align
        else:
            ec = ws.cell(row=row, column=3, value=_money_str(exposure))
            ec.alignment = _align
        row += 1
    # Blank rows for Projected Exposure and Recommended
    ws.cell(row=row, column=2, value="Projected Exposure (30d)").font = Font(italic=True)
    pc = ws.cell(row=row, column=3, value="=P7")
    pc.alignment = Alignment(horizontal="center")
    row += 1
    ws.cell(row=row, column=2, value="Recommended").font = Font(bold=True)
# Remove the broken Chargeback Rate % # row (B34 label + C34 #DIV/0! formula)
    ws["B34"].value = None
    ws["C34"].value = None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
