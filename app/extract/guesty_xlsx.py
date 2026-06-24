"""
Guesty 'Transaction Activity Search Results' XLSX extractor.

Handles the spreadsheet export that the AI/PDF path cannot read:
  - data sheet identified by presence of 'Amount' + 'Type' headers
    (skips the trailing empty 'Pivot Table 1' sheet)
  - date taken from 'Created date (UTC)' (newer export) or 'Date' (legacy)
  - transaction kind read from 'Type' (charge/refund/chargeback);
    falls back to 'Category' and a positive-amount heuristic for legacy files
  - trailing blank columns ignored

Returns the same contract as the other extractors:
  {"table": {month: {sales, refunds, chargebacks}}, "latest_date": datetime}
"""
from collections import defaultdict
from datetime import datetime

import openpyxl


def _month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _to_float(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "").replace("$", "")
    if s == "" or s.lower() in ("null", "na", "nan"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pick_data_sheet(wb):
    for ws in wb.worksheets:
        header = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c).strip() if c is not None else "" for c in row]
            break
        if header and "Amount" in header and "Type" in header:
            return ws, header
    return None, None


def _row_year_month(row_dict):
    """Derive (year, month) from the date column; fall back to Month col + current year."""
    dt = None
    raw_date = row_dict.get("Created date (UTC)")
    if raw_date is None:
        raw_date = row_dict.get("Date")
    if raw_date is not None:
        if isinstance(raw_date, datetime):
            dt = raw_date
        else:
            for fmt in ("%b %d, %Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(str(raw_date).strip(), fmt)
                    break
                except ValueError:
                    continue
    if dt is not None:
        return dt.year, dt.month, dt
    m = row_dict.get("Month")
    if m is not None:
        try:
            mm = int(m)
            yr = datetime.utcnow().year
            return yr, mm, datetime(yr, mm, 1)
        except (ValueError, TypeError):
            pass
    return None, None, None


def extract_guesty_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws, header = _pick_data_sheet(wb)
    if ws is None:
        raise ValueError("No Transaction Activity data sheet (Amount/Type headers) found in workbook")

    col_idx = {name: i for i, name in enumerate(header) if name}

    table = defaultdict(lambda: {"sales": 0.0, "refunds": 0.0, "chargebacks": 0.0})
    latest_date = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        row_dict = {name: row[i] for name, i in col_idx.items() if i < len(row)}
        amt_raw = row_dict.get("Amount")
        if amt_raw is None or str(amt_raw).strip() == "":
            continue
        amt = _to_float(amt_raw)
        if amt == 0.0:
            continue

        year, month, dt = _row_year_month(row_dict)
        if year is None:
            continue
        if dt is not None and (latest_date is None or dt > latest_date):
            latest_date = dt
        mk = _month_key(year, month)

        type_str = str(row_dict.get("Type", "")).lower()
        category = str(row_dict.get("Category", "")).lower()
        tag = type_str + " " + category
        if "chargeback" in tag or "chb" in tag:
            table[mk]["chargebacks"] += abs(amt)
        elif "refund" in tag:
            table[mk]["refunds"] += abs(amt)
        elif "charge" in type_str:
            table[mk]["sales"] += abs(amt)
        elif amt > 0:
            table[mk]["refunds"] += abs(amt)
        else:
            table[mk]["sales"] += abs(amt)

    if not table:
        raise ValueError("Transaction Activity sheet contained no usable rows")

    result = {m: {k: round(v, 2) for k, v in d.items()}
              for m, d in sorted(table.items())}
    return {"table": result, "latest_date": latest_date}
